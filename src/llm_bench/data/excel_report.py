"""Per-run Excel summary for Ollama models only.

The Excel attachment is INTENTIONALLY narrower than the JSON report:
it answers "what was actually loaded on the box, did it support
thinking, and how fast did it serve?" for every Ollama model that
made it past readiness. vLLM / OpenAI rows do NOT appear here — the
Excel exists so an operator can scan a single grid of GPU/RAM/context
/ tokens-per-second figures without filtering on api_type first.

Column layout (kept stable; mirrored by the column header row):

    App                     spec.app_name
    Model                   spec.model_name (server-discovered when applicable)
    API                     always "ollama" in this sheet
    Supports Thinking       runtime probe via /api/show capabilities[]
                            (also drives whether the streaming TTFT
                            probe runs — see orchestrator._run_one_prompt)
    Family                  /api/show details.family
    Parameter Size          /api/show details.parameter_size
    Quantization            /api/show details.quantization_level
    Max Context             /api/show model_info[*.context_length]
    Runtime Context         /api/ps entry.context_length
    Disk (GiB)              /api/tags entry.size / 1 GiB
    Total VRAM+RAM (GiB)    /api/ps entry.size / 1 GiB
    VRAM (GiB)              /api/ps entry.size_vram / 1 GiB
    RAM (GiB)               (size - size_vram) / 1 GiB
    KV Cache (GiB)          (size - disk) / 1 GiB (negative when unloaded)
    Processor Split         "100% GPU" / "X% GPU / Y% CPU" / "not loaded"
    Loaded                  bool — /api/ps had an entry for this model
    Prompts OK              "ok_count / total_count" (across all prompts)
    Avg TTFT (s)            mean of QuestionResult.ttft_seconds (ok rows only)
    Avg Think TTFT (s)      mean of QuestionResult.thinking_ttft_seconds
    Avg TPS                 mean of QuestionResult.tps
    Avg Tokens              mean of QuestionResult.eval_count
    Avg Wall (s)            mean of QuestionResult.wall_seconds
    Total Server (s)        sum of QuestionResult.total_server_seconds
    Install Decision        "fresh" / "reused" / "recovered" / ""
    Install (s)             ModelResult.install_seconds
    Uninstall (s)           ModelResult.uninstall_seconds (0 if skipped)
    Started / Finished      UTC ISO timestamps from the run
    Endpoint                base URL the benchmark used
    Error                   ModelResult.error or "" when OK

`render_ollama_excel` returns `(filename, bytes)` so callers can write
to disk AND attach to the email without re-serializing. When no
Ollama row is present (config has only OpenAI / vLLM models, or all
failed early), the function returns `(None, b"")` and the orchestrator
silently drops the attachment.
"""
from __future__ import annotations

import io
import logging
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from llm_bench.constants import LOG_NAMESPACE
from llm_bench.domain import ApiType, ModelResult

log = logging.getLogger(LOG_NAMESPACE)


_COLUMNS: list[tuple[str, str]] = [
    # (header, descriptor_or_result_key)
    # Header text is what shows up in the .xlsx; the second tuple element
    # is the key used by `_row_for` below.
    ("App",                  "app_name"),
    ("Model",                "model"),
    ("API",                  "api_type"),
    ("Supports Thinking",    "ollama_supports_thinking"),
    ("Family",               "family"),
    ("Parameter Size",       "parameter_size"),
    ("Quantization",         "quantization"),
    ("Max Context",          "max_context"),
    ("Runtime Context",      "runtime_context"),
    ("Disk (GiB)",           "disk_gb"),
    ("Total VRAM+RAM (GiB)", "total_gb"),
    ("VRAM (GiB)",           "vram_gb"),
    ("RAM (GiB)",            "ram_gb"),
    ("KV Cache (GiB)",       "kvcache_gb"),
    ("Processor Split",      "processor"),
    ("Loaded",               "loaded"),
    ("Prompts OK",           "prompts_ok"),
    ("Avg TTFT (s)",         "avg_ttft"),
    ("Avg Think TTFT (s)",   "avg_thinking_ttft"),
    ("Avg TPS",              "avg_tps"),
    ("Avg Tokens",           "avg_eval_count"),
    ("Avg Wall (s)",         "avg_wall"),
    ("Total Server (s)",     "total_server"),
    ("Install Decision",     "install_decision"),
    ("Install (s)",          "install_seconds"),
    ("Uninstall (s)",        "uninstall_seconds"),
    ("Started",              "started_at"),
    ("Finished",             "finished_at"),
    ("Endpoint",             "endpoint"),
    ("Error",                "error"),
]


def _descriptor_field(result: ModelResult, key: str) -> Any:
    """Pull `key` out of `result.ollama_descriptor`, tolerating the
    "descriptor wasn't recorded" case (orchestrator step skipped /
    daemon unreachable). Missing -> None, which renders as an empty
    cell in openpyxl.
    """
    d = result.ollama_descriptor
    if not isinstance(d, dict):
        return None
    return d.get(key)


def _row_for(result: ModelResult) -> list[Any]:
    """Project one :class:`ModelResult` into the column order above."""
    ok_count = sum(1 for q in result.questions if q.ok)
    total_count = len(result.questions)
    total_server = sum(
        q.total_server_seconds for q in result.questions if q.ok)

    field_map: dict[str, Any] = {
        "app_name": result.app_name,
        "model": result.model,
        "api_type": str(result.api_type),
        "ollama_supports_thinking": _format_tristate_bool(
            result.ollama_supports_thinking),
        "family": _descriptor_field(result, "family"),
        "parameter_size": _descriptor_field(result, "parameter_size"),
        "quantization": _descriptor_field(result, "quantization"),
        "max_context": _descriptor_field(result, "max_context"),
        "runtime_context": _descriptor_field(result, "runtime_context"),
        "disk_gb": _descriptor_field(result, "disk_gb"),
        "total_gb": _descriptor_field(result, "total_gb"),
        "vram_gb": _descriptor_field(result, "vram_gb"),
        "ram_gb": _descriptor_field(result, "ram_gb"),
        "kvcache_gb": _descriptor_field(result, "kvcache_gb"),
        "processor": _descriptor_field(result, "processor"),
        "loaded": _descriptor_field(result, "loaded"),
        "prompts_ok": f"{ok_count} / {total_count}",
        "avg_ttft": round(result.avg("ttft_seconds"), 3),
        "avg_thinking_ttft": round(result.avg("thinking_ttft_seconds"), 3),
        "avg_tps": round(result.avg("tps"), 2),
        "avg_eval_count": round(result.avg("eval_count"), 1),
        "avg_wall": round(result.avg("wall_seconds"), 3),
        "total_server": round(total_server, 3),
        "install_decision": str(result.install_decision or ""),
        "install_seconds": result.install_seconds,
        "uninstall_seconds": result.uninstall_seconds,
        "started_at": result.started_at,
        "finished_at": result.finished_at,
        "endpoint": result.endpoint,
        "error": result.error or "",
    }
    return [field_map[key] for _, key in _COLUMNS]


def _format_tristate_bool(value: bool | None) -> str:
    """Render Optional[bool] as Yes / No / "" so empty doesn't look like
    a probe that explicitly returned False.
    """
    if value is True:
        return "Yes"
    if value is False:
        return "No"
    return ""


def _style_header(ws) -> None:
    """Bold + light-gray header row + sensible column widths so the
    sheet is legible without manual fiddling."""
    header_font = Font(bold=True, color="333333")
    header_fill = PatternFill(start_color="F2F4F7",
                              end_color="F2F4F7",
                              fill_type="solid")
    header_align = Alignment(horizontal="left", vertical="center")
    for col_idx, (header, _) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        # Width heuristic: 1.1x the header text, with a 12 / 42 floor + ceiling.
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            12, min(42, int(len(header) * 1.6) + 2))
    ws.freeze_panes = "A2"


def render_ollama_excel(results: list[ModelResult],
                        ) -> tuple[str | None, bytes]:
    """Build the Ollama-only summary workbook.

    Returns ``(filename_hint, content_bytes)``. The filename hint is
    just the basename ``"llm_bench_ollama.xlsx"`` so the caller can
    decide where on disk to write it; bytes are the workbook content
    ready for both ``open(..., "wb").write(...)`` and SMTP attachment.

    When no Ollama row is present, returns ``(None, b"")`` and logs an
    INFO so the caller can drop the attachment silently.
    """
    ollama = [r for r in results if r.api_type == ApiType.OLLAMA]
    if not ollama:
        log.info("excel_report: no ollama results to write; "
                 "skipping .xlsx attachment")
        return None, b""

    wb = Workbook()
    ws = wb.active
    ws.title = "ollama"
    ws.append([header for header, _ in _COLUMNS])
    _style_header(ws)
    for result in ollama:
        ws.append(_row_for(result))

    buf = io.BytesIO()
    wb.save(buf)
    payload = buf.getvalue()
    log.info("excel_report: rendered %d ollama row(s), %d bytes",
             len(ollama), len(payload))
    return "llm_bench_ollama.xlsx", payload


__all__ = ["render_ollama_excel"]
