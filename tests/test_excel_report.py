"""Smoke tests for ``llm_bench.data.excel_report``.

The renderer is pure data-in / bytes-out, so we round-trip the bytes
back through ``openpyxl`` and verify a few representative cells. We
cover three cases:

* mixed Ollama + OpenAI results -> only Ollama rows appear
* no Ollama results              -> empty workbook (filename=None,
                                    bytes=b"")
* Ollama row without a descriptor -> probe / capability columns
                                     render as "" without crashing
"""
from __future__ import annotations

import io

from openpyxl import load_workbook

from llm_bench.data.excel_report import render_ollama_excel
from llm_bench.domain import ApiType, ModelResult, QuestionResult


def _make_ollama_result(*, with_descriptor: bool = True) -> ModelResult:
    """Two-prompt ollama result with one OK row, so the avg() helpers
    have something to chew on.
    """
    return ModelResult(
        app_name="ollama-qwen3",
        model="qwen3:8b",
        api_type=ApiType.OLLAMA,
        started_at="2026-05-12T10:00:00Z",
        finished_at="2026-05-12T10:02:00Z",
        endpoint="https://ollama.example/api",
        install_seconds=42.0,
        uninstall_seconds=7.0,
        ollama_supports_thinking=True,
        ollama_descriptor={
            "model": "qwen3:8b",
            "family": "qwen3",
            "parameter_size": "8.2B",
            "quantization": "Q4_K_M",
            "max_context": 40960,
            "runtime_context": 4096,
            "disk_gb": 4.5,
            "total_gb": 5.1,
            "vram_gb": 5.1,
            "ram_gb": 0.0,
            "kvcache_gb": 0.6,
            "processor": "100% GPU",
            "loaded": True,
        } if with_descriptor else None,
        questions=[
            QuestionResult(
                prompt="hi",
                ok=True,
                wall_seconds=1.0,
                ttft_seconds=0.2,
                thinking_ttft_seconds=0.05,
                eval_count=120,
                tps=85.0,
                has_thinking=True,
                total_server_seconds=0.95,
            ),
            QuestionResult(prompt="boom", ok=False, error="net"),
        ],
    )


def _make_openai_result() -> ModelResult:
    """vLLM-served openai-compatible result. Should be filtered out."""
    return ModelResult(
        app_name="vllm-llama",
        model="meta-llama/Llama-3-8B",
        api_type=ApiType.OPENAI,
        questions=[QuestionResult(prompt="x", ok=True)],
    )


def _load_rows(payload: bytes) -> list[list]:
    """Round-trip the rendered bytes back through openpyxl so we can
    assert on cell values without depending on private internals.
    """
    wb = load_workbook(io.BytesIO(payload), read_only=True)
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


class TestRenderOllamaExcel:
    def test_empty_when_no_ollama_results(self) -> None:
        # OpenAI-only run should produce no workbook at all.
        filename, payload = render_ollama_excel([_make_openai_result()])
        assert filename is None
        assert payload == b""

    def test_filters_to_ollama_only(self) -> None:
        # Mixed list — only the Ollama row survives in the workbook.
        filename, payload = render_ollama_excel([
            _make_openai_result(),
            _make_ollama_result(),
        ])
        assert filename == "llm_bench_ollama.xlsx"
        rows = _load_rows(payload)
        # Header + 1 ollama row.
        assert len(rows) == 2
        header = rows[0]
        assert "App" in header and "Family" in header

        ollama_row = dict(zip(header, rows[1], strict=True))
        assert ollama_row["App"] == "ollama-qwen3"
        assert ollama_row["Model"] == "qwen3:8b"
        assert ollama_row["API"] == "ollama"
        # Runtime probe is the sole thinking signal in the workbook.
        assert ollama_row["Supports Thinking"] == "Yes"
        # Defensive: the dropped redundant columns must NOT come back.
        assert "spec.thinking" not in header
        assert "Has Think (config)" not in header
        # Descriptor fields
        assert ollama_row["Family"] == "qwen3"
        assert ollama_row["Parameter Size"] == "8.2B"
        assert ollama_row["Quantization"] == "Q4_K_M"
        assert ollama_row["Max Context"] == 40960
        assert ollama_row["Processor Split"] == "100% GPU"
        assert ollama_row["Loaded"] is True
        # Aggregated benchmark stats (only the OK row contributes).
        assert ollama_row["Prompts OK"] == "1 / 2"
        assert ollama_row["Avg TPS"] == 85.0
        assert ollama_row["Avg Tokens"] == 120.0

    def test_descriptor_absent_renders_blank_cells(self) -> None:
        result = _make_ollama_result(with_descriptor=False)
        filename, payload = render_ollama_excel([result])
        assert filename is not None
        rows = _load_rows(payload)
        header = rows[0]
        ollama_row = dict(zip(header, rows[1], strict=True))
        # Descriptor-derived fields fall back to empty cells (None ->
        # blank in openpyxl) but the row itself is still emitted.
        assert ollama_row["Family"] in (None, "")
        assert ollama_row["Parameter Size"] in (None, "")
        assert ollama_row["Loaded"] in (None, "")
        # Aggregated benchmark stats still computed from .questions[].
        assert ollama_row["Prompts OK"] == "1 / 2"
